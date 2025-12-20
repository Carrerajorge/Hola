from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, date
from typing import Any, Dict, Iterable, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.cell import (
    coordinate_from_string,
    column_index_from_string,
    get_column_letter,
)
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

from .schemas import ExcelSpec, SheetSpec, TableSpec, ChartSpec


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGN = Alignment(vertical="center", horizontal="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", horizontal="left", wrap_text=True)


def _cell_to_rc(cell: str) -> Tuple[int, int]:
    col, row = coordinate_from_string(cell)
    return int(row), column_index_from_string(col)


def _rc_to_cell(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def _maybe_parse_date(v: Any) -> Any:
    # Accept ISO date strings and coerce to python date/datetime for Excel
    if isinstance(v, str):
        # YYYY-MM-DD
        try:
            if len(v) == 10 and v[4] == "-" and v[7] == "-":
                return date.fromisoformat(v)
        except Exception:
            return v
        # YYYY-MM-DDTHH:MM:SS...
        try:
            if "T" in v and v[:4].isdigit():
                return datetime.fromisoformat(v.replace("Z", "+00:00"))
        except Exception:
            return v
    return v


def _auto_fit_columns(ws, min_width: float = 8.0, max_width: float = 60.0) -> None:
    # Simple heuristic: measure string length of displayed values
    col_max: Dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for j, val in enumerate(row, start=1):
            if val is None:
                continue
            s = str(val)
            col_max[j] = max(col_max.get(j, 0), len(s))
    for j, maxlen in col_max.items():
        # +2 for padding
        width = max(min_width, min(max_width, maxlen + 2))
        ws.column_dimensions[get_column_letter(j)].width = float(width)


def _write_table(ws, table: TableSpec, *, table_index: int) -> Tuple[int, int, int, int]:
    start_r, start_c = _cell_to_rc(table.anchor)

    # headers
    for j, h in enumerate(table.headers):
        cell = ws.cell(row=start_r, column=start_c + j, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    # rows
    for i, row in enumerate(table.rows):
        for j, val in enumerate(row):
            v = _maybe_parse_date(val)
            cell = ws.cell(row=start_r + 1 + i, column=start_c + j, value=v)
            cell.alignment = CELL_ALIGN

    end_r = start_r + len(table.rows)
    end_c = start_c + len(table.headers) - 1

    # formats by header
    header_to_col = {h: start_c + idx for idx, h in enumerate(table.headers)}
    for header, fmt in (table.column_formats or {}).items():
        if header not in header_to_col:
            continue
        c = header_to_col[header]
        for r in range(start_r + 1, end_r + 1):
            ws.cell(row=r, column=c).number_format = fmt

    # Create an Excel Table object (gives styling + autofilter)
    ref = f"{_rc_to_cell(start_r, start_c)}:{_rc_to_cell(end_r, end_c)}"
    t = Table(displayName=f"Table{table_index}", ref=ref)

    style = TableStyleInfo(
        name=table.table_style or "TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    t.tableStyleInfo = style
    ws.add_table(t)

    # Freeze header row if requested (unless worksheet layout overrides later)
    if table.freeze_header:
        ws.freeze_panes = _rc_to_cell(start_r + 1, start_c)

    return start_r, start_c, end_r, end_c


def _add_chart(ws, ch: ChartSpec) -> None:
    if ch.type == "bar":
        chart = BarChart()
    elif ch.type == "line":
        chart = LineChart()
    else:
        chart = PieChart()

    chart.title = ch.title or ""

    cats = Reference(ws, range_string=f"{ws.title}!{ch.categories_range}")
    vals = Reference(ws, range_string=f"{ws.title}!{ch.values_range}")

    if ch.type in ("bar", "line"):
        chart.add_data(vals, titles_from_data=False)
        chart.set_categories(cats)
    else:  # pie
        chart.add_data(vals, titles_from_data=False)
        chart.set_categories(cats)

    ws.add_chart(chart, ch.position)


def render_excel(spec: ExcelSpec, out_path: str) -> str:
    """Render an ExcelSpec into a .xlsx file."""
    wb = Workbook()
    # remove default sheet
    if wb.worksheets:
        wb.remove(wb.worksheets[0])

    for sheet in spec.sheets:
        _render_sheet(wb, sheet)

    wb.properties.title = spec.workbook_title or "Report"
    wb.save(out_path)
    return out_path


def _render_sheet(wb: Workbook, sheet: SheetSpec) -> None:
    ws = wb.create_sheet(title=sheet.name)

    # tables
    for idx, table in enumerate(sheet.tables, start=1):
        _write_table(ws, table, table_index=idx)

    # sheet-level layout overrides
    if sheet.layout.freeze_panes:
        ws.freeze_panes = sheet.layout.freeze_panes

    ws.sheet_view.showGridLines = bool(sheet.layout.show_gridlines)

    # explicit widths
    for col_letter, width in (sheet.layout.column_widths or {}).items():
        ws.column_dimensions[col_letter].width = float(width)

    # auto-fit
    if sheet.layout.auto_fit_columns and not sheet.layout.column_widths:
        _auto_fit_columns(ws)

    # charts
    for ch in sheet.charts:
        _add_chart(ws, ch)
