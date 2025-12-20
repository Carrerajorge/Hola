from __future__ import annotations

from typing import List, Tuple, Union

from openpyxl import load_workbook
from docx import Document

from .schemas import ExcelSpec, DocSpec


def validate_excel_file(path: str) -> List[str]:
    errors: List[str] = []
    try:
        wb = load_workbook(path)
        if not wb.sheetnames:
            errors.append("Workbook has no sheets.")
    except Exception as e:
        errors.append(f"Failed to open xlsx: {e}")
    return errors


def validate_docx_file(path: str) -> List[str]:
    errors: List[str] = []
    try:
        doc = Document(path)
        # basic sanity check
        _ = len(doc.paragraphs)
    except Exception as e:
        errors.append(f"Failed to open docx: {e}")
    return errors


def parse_excel_spec(payload: dict) -> ExcelSpec:
    return ExcelSpec.model_validate(payload)


def parse_doc_spec(payload: dict) -> DocSpec:
    return DocSpec.model_validate(payload)
